# coding = utf8
import os
from time import sleep

import openpyxl
import pandas as pd
import pyautogui

pyautogui.FAILSAFE = True
os.path.abspath(".")
"""
    @Project:pythonProject_seevision
    @File:W3_picture_compare.py
    @Author:十二点前要睡觉
    @Date:2022/4/18 17:29
"""


def openSearchBox(name="D15", zoom_times=6):
    pyautogui.press("s")
    pyautogui.press("s")
    pyautogui.typewrite(" {}".format(name))
    pyautogui.press("enter")
    pyautogui.hotkey("ctrl", "w")
    for i in range(zoom_times):
        pyautogui.doubleClick()
    pyautogui.hotkey("ctrl", "w")


def resize_home():
    pyautogui.press("HOME")


def catchFramePicture(name, e_type, e_degree):
    # C13 _horizontal (0,180)
    # 0402 zoom_times = 6
    e0402_gx_horizontal = screenX * 0.29
    e0402_gy_horizontal = screenY * 0.25
    # vertical (90, 270)  264 * 558 zoom_times = 6
    e0402_gx_vertical = screenX * 0.14
    e0402_gy_vertical = screenY * 0.5

    # C20 _horizontal
    # 0603 zoom_times = 6
    e0603_gx_horizontal = screenX * 0.42
    e0603_gy_horizontal = screenY * 0.28
    # 305 * 824 zoom_times = 6
    e0603_gx_vertical = screenX * 0.17
    e0603_gy_vertical = screenY * 0.76

    # R42 _horizontal
    # 0805 zoom_times = 6
    e0805_gx_horizontal = screenX * 0.64
    e0805_gy_horizontal = screenY * 0.55
    # zoom_times = 5 383*773
    e0805_gx_vertical = screenX * 0.2
    e0805_gy_vertical = screenY * 0.72

    # D17 _horizontal
    # 0603-2 zoom_times = 6
    e0201_gx_horizontal = screenX * 0.22
    e0201_gy_horizontal = screenY * 0.21
    # 229 * 422 zoom_times = 6
    e0201_gx_vertical = screenX * 0.14
    e0201_gy_vertical = screenY * 0.42

    if "0402" in e_type:
        if e_degree == 0 or e_degree == 180:
            GoodX = e0402_gx_horizontal
            GoodY = e0402_gy_horizontal
        elif e_degree == 90 or e_degree == 270:
            GoodX = e0402_gx_vertical
            GoodY = e0402_gy_vertical
        zoomtimes = 6
    elif "0603" in e_type:
        if e_degree == 0 or e_degree == 180:
            GoodX = e0603_gx_horizontal
            GoodY = e0603_gy_horizontal
        elif e_degree == 90 or e_degree == 270:
            GoodX = e0603_gx_vertical
            GoodY = e0603_gy_vertical
        zoomtimes = 6
    elif "0805" in e_type:
        if e_degree == 0 or e_degree == 180:
            GoodX = e0805_gx_horizontal
            GoodY = e0805_gy_horizontal
            zoomtimes = 6
        elif e_degree == 90 or e_degree == 270:
            GoodX = e0805_gx_vertical
            GoodY = e0805_gy_vertical
            zoomtimes = 5
    elif "0603-2" in e_type:
        if e_degree == 0 or e_degree == 180:
            GoodX = e0201_gx_horizontal
            GoodY = e0201_gy_horizontal
        elif e_degree == 90 or e_degree == 270:
            GoodX = e0201_gx_vertical
            GoodY = e0201_gy_vertical
        zoomtimes = 6
    else:
        return "Skip"
    openSearchBox(name, zoomtimes)
    if not os.path.exists("./screenshot/"):
        os.mkdir("./screenshot/")
    x, y = pyautogui.position()
    imagePath = "./screenshot/{}.jpeg".format(name)
    catch_x, catch_y = x - GoodX / 2, y - GoodY / 2
    catch_w, catch_h = GoodX, GoodY
    pyautogui.screenshot(imagePath, region=(catch_x, catch_y, catch_w, catch_h))
    resize_home()
    return "已截取，待核对\n {}".format(imagePath)


"""
    test form read - excel operate
"""


# 从excel中读取数据并返回（element）

def read_excel_for_page_element(form="./sytj0101/工作簿1.xlsx", sheet_name="Sheet1"):
    """
    通过Pandas模块读取case测试点内容，用于后续遍历case执行测试
    :param form:待读取case Excel文件路径
    :param sheet_name:待读取case Excel文件指定sheet表名
    :return:返回对应case所在行以及对应改行case测试点
    """
    print("从excel中读取数据（测试数据case）并返回（element）")
    df = pd.read_excel(form, sheet_name=sheet_name, index_col="number", engine="openpyxl")
    test_case_list = []
    for i in range(1, df.shape[0] + 1):
        part_type = df.loc[i, "PartDecal"]
        ref_des = df.loc[i, "RefDes"]
        layer = df.loc[i, "Layer"]
        orientation = df.loc[i, "Orient."]
        test_case_list.append([i, part_type, ref_des, layer, orientation])
    return test_case_list


def write_into_excel(form="./sytj0101/工作簿1.xlsx", sheet_name="Sheet1", row=1, column=12, value="PASS"):
    """
    通过openpyxl模块将每一行case的测试结果写入对应每一行的结果列中
    :param form:待写入case Excel文件路径
    :param sheet_name:待写入case Excel文件指定sheet表名
    :param row:待写入case测试结果所在行
    :param column:待写入case测试结果所在列
    :param value:待写入测试结果
    :return:None
    """
    print("将测试结果写入excel表格对应Case的行 - 测试结果处：【{}】".format(value))
    wb = openpyxl.load_workbook(form)
    ws = wb[sheet_name]
    grid_value = ws.cell(row + 1, column).value
    print("grid value is {}".format(grid_value))
    if grid_value is None:
        ws.cell(row + 1, column).value = value
    wb.save(form)


if __name__ == '__main__':
    sleep(5)
    element_list = read_excel_for_page_element()
    print(element_list)
    screenX, screenY = pyautogui.size()

    # 手动先设置显示top 或者 bottom的内容
    testSide = "Top"
    for element in element_list:
        e_row = element[0]
        e_type = element[1]
        e_name = element[2]
        e_topOrbottom = element[3]
        e_degree = element[4]
        # 只显示top或bottom的内容
        print(e_row, e_type, e_name, e_topOrbottom)
        if e_topOrbottom in testSide:
            # 增加判断，当是Top，就打开top的显示，是bottom就打开bottom的显示
            # top： 只要参考编号 + 顶面 + 底面
            # 1、Top
            # 2、Silkscreen Top
            # 3、Assembly Drawing Top

            # bottom：
            # 1、Bottom
            # 2、Silkscreen Bottom
            # 3、Assembly Drawing Bottom
            result = catchFramePicture(e_name, e_type, e_degree)
            if result == "Skip":
                pass
            else:
                write_into_excel(form="./sytj0101/工作簿1.xlsx", sheet_name="Sheet1", row=e_row, column=12, value=result)
        else:
            write_into_excel(form="./sytj0101/工作簿1.xlsx", sheet_name="Sheet1", row=e_row, column=12, value="")
        # if e_row == 10:
        #     break
